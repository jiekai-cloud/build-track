import openpyxl
import os

file_path = "生活品質_報價單-案件名稱.xlsx"
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    # Print first 20 rows to analyze structure
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= 20: break
        # Filter out None values for cleaner output
        row_data = [str(cell) if cell is not None else "" for cell in row]
        if any(row_data): # Only print non-empty rows
            print(f"Row {i+1}: {row_data}")
